"""
handlers/xlsx.py
Extracts text from .xlsx files using openpyxl.
Returns a dict mapping sheet name -> text (tab-separated rows).
The top-level 'text' key will be a JSON string of that dict for
easy downstream consumption, but the raw dict is also in metadata.
"""

import json
from pathlib import Path
from typing import Any


def extract(filepath: Path) -> dict[str, Any]:
    try:
        import openpyxl
    except ImportError:
        return {
            "text": None,
            "metadata": {},
            "error": "openpyxl not installed. Run: pip install openpyxl",
        }

    try:
        wb = openpyxl.load_workbook(str(filepath), read_only=True, data_only=True)
    except Exception as exc:
        return {
            "text": None,
            "metadata": {},
            "error": f"Failed to open XLSX: {exc}",
        }

    sheets: dict[str, str] = {}
    sheet_shapes: dict[str, dict[str, int]] = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows_text: list[str] = []
        row_count = 0
        col_count = 0

        for row in ws.iter_rows(values_only=True):
            # Skip entirely empty rows
            if all(cell is None for cell in row):
                continue
            cells = [str(cell) if cell is not None else "" for cell in row]
            rows_text.append("\t".join(cells))
            row_count += 1
            col_count = max(col_count, len(cells))

        sheets[sheet_name] = "\n".join(rows_text)
        sheet_shapes[sheet_name] = {"rows": row_count, "cols": col_count}

    wb.close()

    metadata = {
        "sheet_names": wb.sheetnames,
        "sheet_count": len(wb.sheetnames),
        "sheet_shapes": sheet_shapes,
    }

    return {
        "text": sheets,          # dict[sheet_name, text] as agreed
        "metadata": metadata,
        "error": None,
    }
